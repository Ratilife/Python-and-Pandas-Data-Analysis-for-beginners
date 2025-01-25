#шпаргалка по библиотеке openpyxl, 
'''
Шпаргалка по openpyxl
Установка

pip install openpyxl

Импорт библиотеки
'''
import openpyxl

#Создание новой книги Excel

# Создание новой книги
workbook = openpyxl.Workbook()

# Создание нового листа
    sheet = workbook.active
    sheet.title = "Sheet1"

#Запись данных в ячейки

# Запись данных в ячейки
sheet['A1'] = "Hello"
sheet['B1'] = "World"
sheet['A2'] = 42



# Сохранение книги
workbook.save("example.xlsx")

#Открытие существующей книги

# Открытие существующей книги
workbook = openpyxl.load_workbook("example.xlsx")
sheet = workbook.active  # или workbook['Sheet1'] для конкретного листа



# Чтение данных из ячеек
value = sheet['A1'].value
print(value)  # Вывод: Hello

#Итерация по строкам и столбцам

# Итерация по строкам
for row in sheet.iter_rows(values_only=True):
    print(row)

# Итерация по столбцам
for column in sheet.iter_cols(values_only=True):
    print(column)

#Форматирование ячеек

from openpyxl.styles import Font, Color, Alignment

# Применение стиля к ячейке
sheet['A1'].font = Font(bold=True, color="FF0000")  # Жирный шрифт, красный цвет
sheet['A1'].alignment = Alignment(horizontal='center')  # Центрирование текста



# Добавление формулы
sheet['C1'] = "=SUM(A2:A10)"  # Сумма значений в диапазоне A2:A10

#Добавление изображений

from openpyxl.drawing.image import Image

# Добавление изображения
img = Image("image.png")
sheet.add_image(img, 'D1')  # Добавление изображения в ячейку D1



# Удаление листа
workbook.remove(sheet)



# Закрытие книги (необязательно, но хорошая практика)
workbook.close()
'''
Примечания
openpyxl поддерживает работу с файлами формата .xlsx.
Для работы с файлами формата .xls используйте библиотеку xlrd или xlwt.
Обязательно проверяйте наличие необходимых библиотек и их версии, чтобы избежать ошибок.

Эта шпаргалка охватывает основные функции библиотеки openpyxl. 
'''